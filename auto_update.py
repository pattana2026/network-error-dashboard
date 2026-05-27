"""
auto_update.py
Network Error Dashboard — Auto Update Script
สร้างโดย: Pattana Chancheam

รันอัตโนมัติทุก 08:30 และ 16:30 ผ่าน Windows Task Scheduler
อ่านไฟล์ Excel จาก Network Share → แปลงเป็น data.json
"""

import os
import json
import glob
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl")
    import openpyxl

# ══════════════════════════════════════════════
#  CONFIG — แก้ไขตรงนี้
# ══════════════════════════════════════════════

# โฟลเดอร์ที่เก็บไฟล์ทั้งหมด
OUTPUT_DIR = r"D:\Network Error Dashboard"

# Sheet Name
SHEET_NAME = "Data"  # Sheet ใน atg_site_user.xlsx

# Output
OUTPUT_JSON = os.path.join(OUTPUT_DIR, "data.json")
LOG_FILE    = os.path.join(OUTPUT_DIR, "auto_update.log")

# ══════════════════════════════════════════════

# ══════════════════════════════════════════════
#  EMAIL CONFIG — แก้ไขตรงนี้
# ══════════════════════════════════════════════
EMAIL_ENABLED   = True
EMAIL_SENDER    = "your_email@company.com"       # ← ใส่ Email ผู้ส่ง
EMAIL_PASSWORD  = "your_password"                 # ← ใส่ Password (หรือ App Password)
EMAIL_RECIPIENTS = [
    "person1@company.com",
    "person2@company.com",
    "person3@company.com",
    "person4@company.com",
    "person5@company.com",
]
EMAIL_CC        = ["manager@company.com"]          # ← ใส่ CC
EMAIL_SMTP      = "smtp.office365.com"
EMAIL_PORT      = 587
EMAIL_SEND_TIME = "16:30"                          # ส่งเวลานี้เท่านั้น

def log(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")

def find_latest_excel(folder):
    """หาไฟล์ atg_site_user.xlsx"""
    target = os.path.join(folder, "atg_site_user.xlsx")
    if os.path.exists(target):
        return target
    log("ERROR: ไม่พบไฟล์ atg_site_user.xlsx")
    return None

def read_excel(filepath):
    """อ่าน Excel Sheet Add_Data_Network"""
    log(f"อ่านไฟล์: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # หา Sheet
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        # ถ้าไม่เจอชื่อตรง ใช้ Sheet แรก
        ws = wb.active
        log(f"ไม่พบ Sheet '{SHEET_NAME}' ใช้ Sheet: {ws.title}")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        log("ไม่มีข้อมูลในไฟล์")
        return []

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    log(f"Headers: {headers}")
    log(f"จำนวนแถว: {len(rows)-1}")

    # Map column names to standard Thai keys
    # รองรับทั้งชื่อ English และ Thai
    COL_MAP = {
        # English
        'Station ID':     'รหัสสถานี',
        'Name':           'ชื่อ',
        'Last Inventory': 'คงเหลือล่าสุด',
        'Last Closed Day':'ปิดวันล่าสุด',
        'Remark':         'หมายเหตุ',
        'Province':       'จังหวัด',
        'Phone':          'โทร',
        'Contact':        'ชื่อผู้ติดต่อ',
        # Thai (ถ้าไฟล์มีหัว Thai อยู่แล้ว)
        'รหัสสถานี':      'รหัสสถานี',
        'ชื่อ':           'ชื่อ',
        'คงเหลือล่าสุด':  'คงเหลือล่าสุด',
        'ปิดวันล่าสุด':   'ปิดวันล่าสุด',
        'หมายเหตุ':       'หมายเหตุ',
        'จังหวัด':        'จังหวัด',
        'โทร':            'โทร',
        'ชื่อผู้ติดต่อ':  'ชื่อผู้ติดต่อ',
    }

    data = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        record = {}
        for i, val in enumerate(row):
            if i >= len(headers):
                continue
            col = headers[i]
            if col not in COL_MAP:
                continue
            thai_key = COL_MAP[col]
            if isinstance(val, datetime):
                record[thai_key] = val.strftime("%Y-%m-%d %H:%M")
            elif val is None:
                record[thai_key] = ""
            else:
                record[thai_key] = str(val).strip()
        if record.get('รหัสสถานี','').strip():
            data.append(record)

    wb.close()
    return data

def save_json(data, filepath, source_file):
    """บันทึก data.json"""
    output = {
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source_file": os.path.basename(source_file),
        "total": len(data),
        "data": data
    }
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    log(f"บันทึก data.json สำเร็จ ({len(data)} รายการ)")

def embed_data_into_html(output, html_path):
    """ฝัง JSON data ลงใน index.html โดยตรง"""
    with open(html_path, "r", encoding="utf-8") as f:
        html = f.read()

    # ensure_ascii=True ทำให้ภาษาไทยเป็น \uXXXX (safe ใน JS)
    json_str = json.dumps(output, ensure_ascii=True)
    # escape </script> ไม่ให้ตัด script tag
    json_str = json_str.replace("</" , "</")

    old_line = "const EMBEDDED_DATA = null; // auto_update.py will replace this line"
    new_line = f"const EMBEDDED_DATA = {json_str}; // auto_update.py will replace this line"
    html = html.replace(old_line, new_line, 1)

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    log(f"ฝังข้อมูลลงใน index.html สำเร็จ ({output['total']} รายการ)")


def send_email_report(data, daily_logs, force=False):
    """ส่งรายงาน Email ทุก 16:30"""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders

    if not EMAIL_ENABLED and not force:
        return
    if not EMAIL_SENDER or EMAIL_SENDER == "your_email@company.com":
        log("EMAIL: ยังไม่ได้ตั้งค่า Email ใน config")
        return

    # ตรวจสอบเวลา — ส่งเฉพาะ 16:30 (ยกเว้น force=True)
    now_time = datetime.now().strftime("%H:%M")
    if not force and now_time != EMAIL_SEND_TIME:
        log(f"EMAIL: ไม่ถึงเวลาส่ง ({now_time} != {EMAIL_SEND_TIME})")
        return

    log("EMAIL: กำลังส่งรายงาน...")

    # สร้างสถิติ
    today = datetime.now().strftime("%Y-%m-%d")
    active = [r for r in data]
    crisis = [r for r in active if r.get('ageDays', 0) > 3]

    # สร้าง Email body
    body = f"""
สรุปรายงาน Network Error Dashboard
วันที่: {datetime.now().strftime("%d/%m/%Y %H:%M")}

🔴 Network Error ทั้งหมด: {len(active)} สถานี
🚨 วิกฤติ (>3 วัน): {len(crisis)} สถานี

"""
    if crisis:
        body += "รายการวิกฤติ:\n"
        for r in sorted(crisis, key=lambda x: x.get('คงเหลือล่าสุด',''))[:10]:
            sid  = r.get('รหัสสถานี','')
            name = r.get('ชื่อ','')
            prov = r.get('จังหวัด','')
            note = r.get('หมายเหตุ','')
            body += f"  - {sid} {name} ({prov})\n"
            if note:
                body += f"    หมายเหตุ: {note}\n" 

    body += "\nดูรายละเอียดเพิ่มเติมได้ที่ Dashboard:\nhttps://pattana2026.github.io/network-error-dashboard/\n" 

    # สร้าง Excel attachment
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()

        # Sheet 1: Network Error
        ws1 = wb.active
        ws1.title = "Network Error"
        headers = ['รหัสสถานี','ชื่อสถานี','จังหวัด','คงเหลือล่าสุด','ปิดวันล่าสุด','หมายเหตุ']
        ws1.append(headers)
        for r in active:
            ws1.append([r.get('รหัสสถานี',''), r.get('ชื่อ',''), r.get('จังหวัด',''),
                        r.get('คงเหลือล่าสุด',''), r.get('ปิดวันล่าสุด',''), r.get('หมายเหตุ','')])

        # Sheet 2: ประวัติรายวัน
        ws2 = wb.create_sheet("ประวัติรายวัน")
        ws2.append(['วันที่','รหัสสถานี','ชื่อสถานี','จังหวัด','คงเหลือล่าสุด','หมายเหตุ'])
        for date in sorted(daily_logs.keys()):
            for s in daily_logs[date]:
                ws2.append([date, s.get('รหัสสถานี',''), s.get('ชื่อ',''),
                            s.get('จังหวัด',''), s.get('คงเหลือล่าสุด',''), s.get('หมายเหตุ','')])

        # Save to temp file
        excel_path = os.path.join(OUTPUT_DIR, f"report_{today}.xlsx")
        wb.save(excel_path)
        log(f"สร้างไฟล์ Excel: {excel_path}")
    except Exception as e:
        log(f"EMAIL: สร้าง Excel ไม่สำเร็จ: {e}")
        excel_path = None

    try:
        msg = MIMEMultipart()
        msg['From']    = EMAIL_SENDER
        msg['To']      = ', '.join(EMAIL_RECIPIENTS)
        msg['CC']      = ', '.join(EMAIL_CC)
        msg['Subject'] = f"[Network Error Dashboard] สรุป {now_time} — {datetime.now().strftime('%d/%m/%Y')}"
        msg.attach(MIMEText(body, 'plain', 'utf-8'))

        # Attach Excel
        if excel_path and os.path.exists(excel_path):
            with open(excel_path, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="network_error_report_{today}.xlsx"')
            msg.attach(part)

        all_recipients = EMAIL_RECIPIENTS + EMAIL_CC
        with smtplib.SMTP(EMAIL_SMTP, EMAIL_PORT) as server:
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.sendmail(EMAIL_SENDER, all_recipients, msg.as_string())

        log(f"EMAIL: ส่งสำเร็จ → {', '.join(EMAIL_RECIPIENTS)}")

        # ลบไฟล์ temp
        if excel_path and os.path.exists(excel_path):
            os.remove(excel_path)

    except Exception as e:
        log(f"EMAIL: ส่งไม่สำเร็จ: {e}")


def main():
    # ── Test Email Mode ──
    if len(sys.argv) > 1 and sys.argv[1] == "--test-email":
        log("=" * 50)
        log("ทดสอบส่ง Email...")
        test_data = [{"รหัสสถานี":"TEST001","ชื่อ":"สถานีทดสอบ","จังหวัด":"กรุงเทพ","คงเหลือล่าสุด":"2026-05-24 08:00","หมายเหตุ":"ทดสอบระบบ Email"}]
        send_email_report(test_data, {}, force=True)
        log("=" * 50)
        return

    log("=" * 50)
    log("Network Error Dashboard — Auto Update เริ่มทำงาน")

    # 1. หาไฟล์ Excel ล่าสุดใน D:\Network Error Dashboard
    log(f"ค้นหาไฟล์ Excel ใน: {OUTPUT_DIR}")
    excel_file = find_latest_excel(OUTPUT_DIR)
    if not excel_file:
        log(f"ERROR: ไม่พบไฟล์ Excel ใน {month_folder}")
        sys.exit(1)

    log(f"ไฟล์ล่าสุด: {os.path.basename(excel_file)}")

    # 3. อ่าน Excel
    try:
        data = read_excel(excel_file)
    except Exception as e:
        log(f"ERROR อ่านไฟล์: {e}")
        sys.exit(1)

    if not data:
        log("ไม่มีข้อมูล")
        sys.exit(1)

    # 4. บันทึก JSON + daily log + ฝังลงใน index.html
    try:
        today_str = datetime.now().strftime("%Y-%m-%d")

        # โหลด daily_logs เดิม (ถ้ามี)
        daily_log_file = os.path.join(OUTPUT_DIR, "daily_logs.json")
        if os.path.exists(daily_log_file):
            with open(daily_log_file, "r", encoding="utf-8") as f:
                daily_logs = json.load(f)
        else:
            daily_logs = {}

        # บันทึก snapshot วันนี้
        daily_logs[today_str] = data
        # เก็บย้อนหลังสูงสุด 90 วัน
        if len(daily_logs) > 90:
            oldest = sorted(daily_logs.keys())[0]
            del daily_logs[oldest]
            log(f"ลบ log เก่า: {oldest}")

        # บันทึก daily_logs.json
        with open(daily_log_file, "w", encoding="utf-8") as f:
            json.dump(daily_logs, f, ensure_ascii=False, indent=2)
        log(f"บันทึก daily_logs.json สำเร็จ ({len(daily_logs)} วัน)")

        output = {
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "source_file": os.path.basename(excel_file),
            "total": len(data),
            "data": data,
            "daily_logs": daily_logs
        }
        with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=2)
        log(f"บันทึก data.json สำเร็จ ({len(data)} รายการ)")

        html_path = os.path.join(OUTPUT_DIR, "index.html")
        embed_data_into_html(output, html_path)
    except Exception as e:
        log(f"ERROR: {e}")
        sys.exit(1)

    log("Auto Update เสร็จสิ้น ✓")

    # 5. Push ขึ้น GitHub
    push_to_github()

    # 6. ส่ง Email (เฉพาะ 16:30)
    send_email_report(data, daily_logs)

    log("=" * 50)

def push_to_github():
    import subprocess
    log("Push ขึ้น GitHub...")
    cmds = [
        ["git", "-C", OUTPUT_DIR, "add", "data.json", "index.html", "daily_logs.json"],
        ["git", "-C", OUTPUT_DIR, "commit", "-m",
         f"Auto update data {datetime.now().strftime('%Y-%m-%d %H:%M')}"],
        ["git", "-C", OUTPUT_DIR, "push", "origin", "main"],
    ]
    for cmd in cmds:
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode != 0:
            # commit อาจ return 1 ถ้าไม่มีการเปลี่ยนแปลง — ไม่ถือว่า error
            if "nothing to commit" in result.stdout + result.stderr:
                log("ไม่มีการเปลี่ยนแปลง ไม่ต้อง commit")
                return
            log(f"Git warning: {result.stderr.strip()}")
        else:
            log(f"OK: {' '.join(cmd[2:])}")

if __name__ == "__main__":
    main()
